from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import List, Optional
import csv
import json
import os
from sqlalchemy.orm import Session
from sqlalchemy import func

from models import (
    Application, Group, PaymentLog, FinancialExport, AppState,
    ReminderLog, PaymentOrder, Refund, BankReconciliation, BankReconciliationItem
)
from schemas import DailyReminderItem, DailyReminderResponse
from services.pricing import calc_balance_deadline


def generate_daily_reminders(db: Session, target_date: date) -> DailyReminderResponse:
    reminders = db.query(Application).filter(
        Application.info_completed == True,
        Application.paid_balance < Application.total_price - Application.paid_deposit,
        Application.state != AppState.CANCELLED
    ).all()

    items = []
    for app in reminders:
        group = db.query(Group).filter(Group.id == app.group_id).first()
        balance_deadline, _, _ = calc_balance_deadline(group.departure_date, date.today())

        item = DailyReminderItem(
            app_id=app.id,
            name=app.name,
            phone=app.phone,
            tour_code=group.code,
            departure=group.departure_date,
            total=app.total_price,
            paid=app.paid_deposit + app.paid_balance,
            balance=app.total_price - app.paid_deposit - app.paid_balance,
            balance_deadline=balance_deadline
        )
        items.append(item)

    return DailyReminderResponse(
        date=target_date,
        items=items,
        count=len(items)
    )


def generate_finance_export(db: Session, target_date: date,
                            fields: Optional[List[str]] = None) -> List[dict]:
    payment_logs = db.query(PaymentLog).filter(
        func.date(PaymentLog.created_at) == target_date
    ).all()

    all_fields = ["payment_id", "application_id", "tour_code", "type", "amount", "payment_method", "created_at"]
    export_fields = fields if fields else all_fields

    records = []
    for log in payment_logs:
        app = db.query(Application).filter(Application.id == log.application_id).first()
        group = db.query(Group).filter(Group.id == app.group_id).first() if app else None

        full_record = {
            "payment_id": log.id,
            "application_id": log.application_id,
            "tour_code": group.code if group else None,
            "type": log.type,
            "amount": float(log.amount),
            "payment_method": log.payment_method,
            "created_at": log.created_at.isoformat()
        }

        record = {k: full_record.get(k) for k in export_fields if k in full_record}
        records.append(record)

    return records


def save_finance_export(db: Session, target_date: date, records: List[dict],
                        fmt: str = "csv") -> FinancialExport:
    export_dir = "exports"
    os.makedirs(export_dir, exist_ok=True)

    if fmt == "csv":
        file_name = f"finance_{target_date.isoformat()}.csv"
        file_path = os.path.join(export_dir, file_name)
        if records:
            fieldnames = list(records[0].keys())
        else:
            fieldnames = ["payment_id", "application_id", "tour_code", "type", "amount", "payment_method", "created_at"]
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(records)
    elif fmt == "json":
        file_name = f"finance_{target_date.isoformat()}.json"
        file_path = os.path.join(export_dir, file_name)
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
    elif fmt == "excel":
        import openpyxl
        from openpyxl.styles import Font, Alignment
        file_name = f"finance_{target_date.isoformat()}.xlsx"
        file_path = os.path.join(export_dir, file_name)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "财务数据"
        if records:
            fieldnames = list(records[0].keys())
        else:
            fieldnames = ["payment_id", "application_id", "tour_code", "type", "amount", "payment_method", "created_at"]
        # 写表头
        header_font = Font(bold=True)
        for col, name in enumerate(fieldnames, 1):
            cell = ws.cell(row=1, column=col, value=name)
            cell.font = header_font
        # 写数据
        for row_idx, record in enumerate(records, 2):
            for col_idx, field in enumerate(fieldnames, 1):
                ws.cell(row=row_idx, column=col_idx, value=record.get(field))
        wb.save(file_path)
    else:
        raise ValueError(f"Unsupported format: {fmt}")

    export = FinancialExport(
        export_date=target_date,
        file_path=file_path,
        file_format=fmt,
        record_count=len(records)
    )
    db.add(export)
    db.commit()
    db.refresh(export)
    return export


def generate_finance_report(db: Session, period: str, start_date: date, end_date: date) -> dict:
    payment_logs = db.query(PaymentLog).filter(
        func.date(PaymentLog.created_at) >= start_date,
        func.date(PaymentLog.created_at) <= end_date
    ).all()

    total_deposits = Decimal("0")
    total_balances = Decimal("0")
    for log in payment_logs:
        if log.type == "deposit":
            total_deposits += log.amount
        elif log.type == "balance":
            total_balances += log.amount

    refunds = db.query(Refund).filter(
        func.date(Refund.refunded_at) >= start_date,
        func.date(Refund.refunded_at) <= end_date,
        Refund.status.in_(["completed", "approved"])
    ).all()

    total_refunds = sum(r.refund_amount for r in refunds)
    net_income = total_deposits + total_balances - total_refunds

    return {
        "period": period,
        "total_deposits": total_deposits,
        "total_balances": total_balances,
        "total_refunds": total_refunds,
        "net_income": net_income,
        "record_count": len(payment_logs),
    }


def log_reminder(db: Session, app_id: int, reminder_type: str, content: Optional[str] = None) -> ReminderLog:
    log = ReminderLog(
        application_id=app_id,
        reminder_type=reminder_type,
        content=content,
    )
    db.add(log)
    db.commit()
    db.refresh(log)
    return log


def generate_flow_trend(db: Session, days: int = 7) -> List[dict]:
    """生成近 N 天流水趋势数据（按日聚合收入/退款）"""
    from sqlalchemy import func as sa_func
    from models import PaymentLog, Refund

    today = date.today()
    result = []

    for i in range(days - 1, -1, -1):
        day = today - timedelta(days=i)
        # 当日收入（deposit + balance）
        income = db.query(sa_func.coalesce(sa_func.sum(PaymentLog.amount), 0)).filter(
            sa_func.date(PaymentLog.created_at) == day
        ).scalar()
        # 当日退款（已完成/已批准的）
        refund = db.query(sa_func.coalesce(sa_func.sum(Refund.refund_amount), 0)).filter(
            sa_func.date(Refund.refunded_at) == day,
            Refund.status.in_(["completed", "approved"])
        ).scalar()
        result.append({
            "date": day.isoformat(),
            "income": float(income),
            "refund": float(refund),
        })
    return result


def generate_batch_documents(db: Session, application_ids: List[int], doc_type: str) -> dict:
    applications = db.query(Application).filter(Application.id.in_(application_ids)).all()
    documents = []

    for app in applications:
        group = db.query(Group).filter(Group.id == app.group_id).first()
        if doc_type == "confirmation":
            doc = {
                "application_id": app.id,
                "name": app.name,
                "tour_code": group.code if group else "",
                "departure_date": group.departure_date.isoformat() if group else "",
                "total_price": float(app.total_price),
                "doc_type": "confirmation",
            }
        else:
            order = db.query(PaymentOrder).filter(
                PaymentOrder.application_id == app.id,
                PaymentOrder.order_type == "payment_order"
            ).first()
            doc = {
                "application_id": app.id,
                "name": app.name,
                "tour_code": group.code if group else "",
                "order_no": order.order_no if order else "",
                "balance": float(app.total_price - app.paid_deposit - app.paid_balance),
                "doc_type": "payment_order",
            }
        documents.append(doc)

    return {
        "documents": documents,
        "total_count": len(documents),
    }


def import_bank_statement(db: Session, file_name: str, records: List[dict]) -> BankReconciliation:
    reconciliation = BankReconciliation(
        import_date=date.today(),
        file_name=file_name,
        total_records=len(records),
    )
    db.add(reconciliation)
    db.flush()

    matched_count = 0
    for rec in records:
        bank_amount = Decimal(str(rec.get("amount", 0)))
        bank_date_str = rec.get("date")
        bank_date = date.fromisoformat(bank_date_str) if bank_date_str else None
        bank_ref = rec.get("reference", "")

        matched_payment = db.query(PaymentLog).filter(
            PaymentLog.amount == bank_amount,
            func.date(PaymentLog.created_at) == bank_date if bank_date else True,
        ).first()

        is_matched = matched_payment is not None
        if is_matched:
            matched_count += 1

        item = BankReconciliationItem(
            reconciliation_id=reconciliation.id,
            bank_date=bank_date,
            bank_amount=bank_amount,
            bank_ref=bank_ref,
            matched_payment_id=matched_payment.id if matched_payment else None,
            is_matched=is_matched,
        )
        db.add(item)

    reconciliation.matched_count = matched_count
    reconciliation.unmatched_count = len(records) - matched_count
    db.commit()
    db.refresh(reconciliation)
    return reconciliation
