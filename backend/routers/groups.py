from typing import List, Optional
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload

from database import get_db
from models import Group, Route, Application, AppState
from schemas import GroupCreate, GroupUpdate, GroupResponse, GroupDetailResponse
from services.pricing import calc_deposit, calc_balance_deadline
from schemas import PricingPreviewResponse, BalanceDeadlineResponse

router = APIRouter(prefix="/api/groups", tags=["groups"])


@router.get("", response_model=List[GroupResponse])
def list_groups(
    status: Optional[str] = Query(None, description="Filter by published status: 'published' or 'all'"),
    route_id: Optional[int] = Query(None),
    departure_from: Optional[date] = Query(None),
    departure_to: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(Group)

    if status == "published":
        query = query.filter(Group.is_published == True)
    elif status is None:
        pass

    if route_id:
        query = query.filter(Group.route_id == route_id)
    if departure_from:
        query = query.filter(Group.departure_date >= departure_from)
    if departure_to:
        query = query.filter(Group.departure_date <= departure_to)

    groups = query.order_by(Group.departure_date).all()
    return groups


@router.get("/template")
def download_group_template():
    import openpyxl
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "旅游团导入模板"
    headers = ["路线ID", "团代码", "出发日期(YYYY-MM-DD)", "截止日期(YYYY-MM-DD)", "最大人数", "成人价", "儿童价", "是否发布(True/False)"]
    ws.append(headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=group_template.xlsx"}
    )


@router.post("/import")
def import_groups_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    import openpyxl
    from io import BytesIO
    from datetime import datetime as dt

    content = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    imported = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row[0] or not row[1]:
            continue
        try:
            route_id = int(row[0])
            route = db.query(Route).filter(Route.id == route_id).first()
            if not route:
                errors.append(f"行{idx}: 路线ID {route_id} 不存在")
                continue

            code = str(row[1])
            departure_date = row[2] if isinstance(row[2], date) else dt.strptime(str(row[2]), "%Y-%m-%d").date()
            deadline = row[3] if isinstance(row[3], date) else dt.strptime(str(row[3]), "%Y-%m-%d").date()

            if deadline >= departure_date:
                errors.append(f"行{idx}: 截止日期必须在出发日期之前")
                continue

            group = Group(
                route_id=route_id,
                code=code,
                departure_date=departure_date,
                deadline=deadline,
                max_pax=int(row[4]) if row[4] else 20,
                adult_price=row[5] if row[5] else None,
                child_price=row[6] if row[6] else None,
                is_published=bool(row[7]) if row[7] is not None else False,
            )
            db.add(group)
            imported += 1
        except Exception as e:
            errors.append(f"行{idx}: {str(e)}")

    db.commit()
    return {"imported": imported, "errors": errors}


@router.get("/{group_id}", response_model=GroupDetailResponse)
def get_group(group_id: int, db: Session = Depends(get_db)):
    group = db.query(Group).options(joinedload(Group.route)).filter(Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    return group


@router.post("", response_model=GroupResponse, status_code=201)
def create_group(data: GroupCreate, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == data.route_id).first()
    if not route:
        raise HTTPException(status_code=400, detail="Route not found")

    if data.deadline >= data.departure_date:
        raise HTTPException(status_code=400, detail="Deadline must be before departure date")

    group = Group(
        route_id=data.route_id,
        code=data.code,
        departure_date=data.departure_date,
        deadline=data.deadline,
        max_pax=data.max_pax,
        adult_price=data.adult_price,
        child_price=data.child_price
    )
    db.add(group)
    db.commit()
    db.refresh(group)
    return group


@router.put("/{group_id}", response_model=GroupResponse)
def update_group(group_id: int, data: GroupUpdate, db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if group.is_published:
        raise HTTPException(status_code=400, detail="Cannot modify published group")

    update_data = data.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(group, field, value)

    db.commit()
    db.refresh(group)
    return group


@router.post("/{group_id}/publish", response_model=GroupResponse)
def publish_group(group_id: int, db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if group.is_published:
        raise HTTPException(status_code=400, detail="Group is already published")

    if group.adult_price is None or group.child_price is None:
        raise HTTPException(status_code=400, detail="Prices must be set before publishing")

    group.is_published = True
    db.commit()
    db.refresh(group)
    return group


@router.get("/{group_id}/availability", response_model=dict)
def check_availability(group_id: int, db: Session = Depends(get_db)):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    total_occupied = db.query(
        Application.adults + Application.children
    ).filter(
        Application.group_id == group_id,
        Application.state != AppState.CANCELLED
    ).all()

    occupied_count = sum(x[0] for x in total_occupied) if total_occupied else 0
    available = group.max_pax - occupied_count

    return {
        "group_id": group_id,
        "max_pax": group.max_pax,
        "occupied": occupied_count,
        "available": available
    }


@router.get("/{group_id}/pricing-preview", response_model=PricingPreviewResponse)
def pricing_preview(
    group_id: int,
    adults: int = Query(..., gt=0),
    children: int = Query(0, ge=0),
    db: Session = Depends(get_db)
):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    adult_price = group.adult_price or 0
    child_price = group.child_price or 0
    deposit, rate_str = calc_deposit(
        group.departure_date, adults, children,
        adult_price, child_price
    )
    total_price = adults * adult_price + children * child_price

    return PricingPreviewResponse(
        deposit=deposit,
        total_price=total_price,
        deposit_rate=rate_str
    )


@router.get("/{group_id}/balance-deadline", response_model=BalanceDeadlineResponse)
def get_balance_deadline(
    group_id: int,
    today: date = Query(None),
    db: Session = Depends(get_db)
):
    group = db.query(Group).filter(Group.id == group_id).first()
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")

    if today is None:
        today = date.today()

    balance_deadline, base_deadline, fallback_deadline = calc_balance_deadline(
        group.departure_date, today
    )

    return BalanceDeadlineResponse(
        balance_deadline=balance_deadline,
        base_deadline=base_deadline,
        fallback_deadline=fallback_deadline
    )
