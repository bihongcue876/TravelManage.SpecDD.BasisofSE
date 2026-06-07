from typing import List, Optional
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_

from database import get_db
from models import Route, RouteHistory
from schemas import RouteCreate, RouteUpdate, RouteResponse


class BatchUpdateRequest(BaseModel):
    ids: List[int]
    is_active: bool


router = APIRouter(prefix="/api/routes", tags=["routes"])


def _generate_route_code(db: Session) -> str:
    """生成路线编号：RT + 10 位数字，顺序递增"""
    last = db.query(Route).order_by(Route.id.desc()).first()
    if last and last.code and last.code.startswith("RT"):
        try:
            last_num = int(last.code[2:])
            return f"RT{last_num + 1:010d}"
        except ValueError:
            pass
    return "RT0000000001"


def _parse_bool(v) -> bool:
    """解析 Excel 中的 TRUE/FALSE 字符串或 bool 值"""
    if isinstance(v, bool):
        return v
    if isinstance(v, str):
        return v.strip().upper() == "TRUE"
    return False


@router.get("", response_model=List[RouteResponse])
def list_routes(db: Session = Depends(get_db)):
    routes = db.query(Route).all()
    return routes


@router.get("/search", response_model=List[RouteResponse])
def search_routes(q: str = Query("", min_length=0), db: Session = Depends(get_db)):
    """按路线名称、编号、描述模糊搜索"""
    if not q:
        return db.query(Route).all()
    query = db.query(Route).filter(
        or_(
            Route.name.ilike(f"%{q}%"),
            Route.code.ilike(f"%{q}%"),
            Route.descr.ilike(f"%{q}%"),
        )
    )
    return query.all()


@router.get("/template")
def download_route_template():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "路线导入模板"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["路线名称", "描述", "状态"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 12

    # 示例行（带删除提示）
    ws.append(["示例路线", "路线描述（上传前删掉本行）", "TRUE"])
    ws.append([])
    ws.append(["← 请从此行开始填写数据，示例行务必删除"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=route_template.xlsx"}
    )


@router.post("/import")
def import_routes_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    import openpyxl
    from io import BytesIO

    content = file.file.read()
    file.file.close()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    # 预先取最后一个编号，后续行在内存中递增，避免批量插入冲突
    last = db.query(Route).order_by(Route.id.desc()).first()
    next_num = 1
    if last and last.code and last.code.startswith("RT"):
        try:
            next_num = int(last.code[2:]) + 1
        except ValueError:
            pass

    imported = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        route = Route(
            code=f"RT{next_num:010d}",
            name=str(row[0]).strip(),
            descr=str(row[1]).strip() if row[1] else None,
            is_active=_parse_bool(row[2]) if len(row) > 2 else True,
        )
        db.add(route)
        next_num += 1
        imported += 1

    db.commit()
    return {"imported": imported}


@router.put("/batch")
def batch_update_routes(data: BatchUpdateRequest, db: Session = Depends(get_db)):
    """批量启用/停用路线"""
    routes = db.query(Route).filter(Route.id.in_(data.ids)).all()
    if not routes:
        raise HTTPException(status_code=404, detail="No routes found")

    for route in routes:
        history = RouteHistory(
            route_id=route.id,
            code=route.code,
            name=route.name,
            descr=route.descr,
            is_active=route.is_active,
        )
        db.add(history)
        route.is_active = data.is_active

    db.commit()
    return {"updated": len(routes)}


@router.delete("/{route_id}/force")
def delete_route_force(route_id: int, db: Session = Depends(get_db)):
    """硬删除路线（仅当无关联旅游团时可删除）"""
    route = db.query(Route).filter(Route.id == route_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")

    if route.groups:
        raise HTTPException(
            status_code=400,
            detail=f"该路线下有 {len(route.groups)} 个关联旅游团，无法删除。请先删除关联旅游团或使用停用功能。"
        )

    db.query(RouteHistory).filter(RouteHistory.route_id == route_id).delete()
    db.delete(route)
    db.commit()
    return {"message": "Route deleted successfully"}


@router.get("/{route_id}", response_model=RouteResponse)
def get_route(route_id: int, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")
    return route


@router.post("", response_model=RouteResponse, status_code=201)
def create_route(data: RouteCreate, db: Session = Depends(get_db)):
    route = Route(
        code=_generate_route_code(db),
        name=data.name,
        descr=data.descr,
        is_active=data.is_active,
    )
    db.add(route)
    db.commit()
    db.refresh(route)
    return route


@router.put("/{route_id}", response_model=RouteResponse)
def update_route(route_id: int, data: RouteUpdate, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")

    history = RouteHistory(
        route_id=route.id,
        code=route.code,
        name=route.name,
        descr=route.descr,
        is_active=route.is_active,
    )
    db.add(history)

    if data.name is not None:
        route.name = data.name
    if data.descr is not None:
        route.descr = data.descr
    if data.is_active is not None:
        route.is_active = data.is_active

    db.commit()
    db.refresh(route)
    return route


@router.delete("/{route_id}", status_code=204)
def deactivate_route(route_id: int, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")

    history = RouteHistory(
        route_id=route.id,
        code=route.code,
        name=route.name,
        descr=route.descr,
        is_active=route.is_active,
    )
    db.add(history)

    route.is_active = False
    db.commit()
    return None


@router.get("/{route_id}/history")
def get_route_history(route_id: int, db: Session = Depends(get_db)):
    route = db.query(Route).filter(Route.id == route_id).first()
    if not route:
        raise HTTPException(status_code=404, detail="Route not found")
    history = db.query(RouteHistory).filter(
        RouteHistory.route_id == route_id
    ).order_by(RouteHistory.changed_at.desc()).all()
    return history
