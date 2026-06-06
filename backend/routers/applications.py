import os
from typing import List, Optional
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session, joinedload

from database import get_db
from models import Application, Participant, AppState, PaymentLog, Refund, ParticipantEditHistory
from schemas import (
    ApplicationCreate, ApplicationResponse, ApplicationDetailResponse,
    ParticipantCreate, ParticipantUpdate, ParticipantResponse,
    PaymentRequest, ParticipantsBulkRequest, CancelPreviewResponse,
    CancelPreviewDetailResponse, ApplicationSearchParams,
    PaymentLogDetailResponse, RemainingBalanceResponse,
    DepositPreviewResponse, ParticipantEditHistoryResponse,
    DuplicateParticipantWarning, RefundDetailResponse,
    PartialCancelRequest, RefundApprovalRequest,
    ReminderLogResponse, PaymentOrderResponse,
)
from services import application as app_service
from services.pricing import calc_balance_deadline

router = APIRouter(prefix="/api/applications", tags=["applications"])


@router.post("", response_model=ApplicationResponse, status_code=201)
def create_application(data: ApplicationCreate, db: Session = Depends(get_db)):
    try:
        application = app_service.create_application(db, data)
        return application
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/search", response_model=List[ApplicationResponse])
def search_applications(
    code: Optional[str] = Query(None),
    departure_from: Optional[date] = Query(None),
    departure_to: Optional[date] = Query(None),
    name: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    query = db.query(Application).join(Application.group)

    if code:
        query = query.filter(Application.group.has(code=code))
    if departure_from:
        query = query.filter(Application.group.has(departure_date=departure_from))
    if departure_to:
        query = query.filter(Application.group.has(departure_date=departure_to))
    if name:
        query = query.filter(Application.name.like(f"%{name}%"))

    applications = query.all()
    return applications


@router.get("/{application_id}", response_model=ApplicationDetailResponse)
def get_application(application_id: int, db: Session = Depends(get_db)):
    application = db.query(Application).options(
        joinedload(Application.participants),
        joinedload(Application.group)
    ).filter(Application.id == application_id).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    return application


@router.get("/{application_id}/deposit-preview", response_model=DepositPreviewResponse)
def get_deposit_preview(
    application_id: int,
    db: Session = Depends(get_db)
):
    application = db.query(Application).filter(Application.id == application_id).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    try:
        preview = app_service.get_deposit_preview(
            db, application.group_id, application.adults, application.children
        )
        return preview
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{application_id}/pay-deposit", response_model=ApplicationResponse)
def pay_deposit(application_id: int, data: PaymentRequest, db: Session = Depends(get_db)):
    try:
        application = app_service.pay_deposit(
            db, application_id, data.amount,
            payment_method=data.payment_method,
            voucher_paths=data.voucher_paths
        )
        return application
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{application_id}/pay-balance", response_model=ApplicationResponse)
def pay_balance(application_id: int, data: PaymentRequest, db: Session = Depends(get_db)):
    try:
        application = app_service.pay_balance(
            db, application_id, data.amount,
            payment_method=data.payment_method,
            voucher_paths=data.voucher_paths
        )
        return application
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{application_id}/remaining-balance", response_model=RemainingBalanceResponse)
def get_remaining_balance(application_id: int, db: Session = Depends(get_db)):
    try:
        result = app_service.get_remaining_balance(db, application_id)
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{application_id}/payment-logs", response_model=List[PaymentLogDetailResponse])
def get_payment_logs(application_id: int, db: Session = Depends(get_db)):
    from models import PaymentVoucher
    logs = db.query(PaymentLog).options(
        joinedload(PaymentLog.vouchers)
    ).filter(PaymentLog.application_id == application_id).all()
    return logs


@router.post("/{application_id}/payment-voucher")
def upload_payment_voucher(
    application_id: int,
    file: UploadFile = File(...),
    payment_log_id: Optional[int] = Query(None),
    db: Session = Depends(get_db)
):
    from models import PaymentVoucher
    upload_dir = "uploads/vouchers"
    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, f"{application_id}_{file.filename}")
    with open(file_path, "wb") as f:
        content = file.file.read()
        f.write(content)

    if payment_log_id:
        voucher = PaymentVoucher(
            payment_log_id=payment_log_id,
            file_name=file.filename,
            file_path=file_path,
            file_size=len(content),
        )
        db.add(voucher)
        db.commit()
        db.refresh(voucher)
        return {"id": voucher.id, "file_name": voucher.file_name, "file_path": voucher.file_path}

    return {"file_name": file.filename, "file_path": file_path}


@router.post("/{application_id}/participants", response_model=ApplicationResponse)
def add_participants(
    application_id: int,
    data: ParticipantsBulkRequest,
    db: Session = Depends(get_db)
):
    try:
        application = app_service.add_participants(db, application_id, data.participants)
        return application
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{application_id}/participants", response_model=List[ParticipantResponse])
def list_participants(application_id: int, db: Session = Depends(get_db)):
    participants = db.query(Participant).filter(
        Participant.application_id == application_id
    ).all()
    return participants


@router.put("/participants/{participant_id}", response_model=ParticipantResponse)
def update_participant(
    participant_id: int,
    data: ParticipantUpdate,
    edited_by: str = Query("admin"),
    db: Session = Depends(get_db)
):
    try:
        participant = app_service.update_participant(db, participant_id, data, edited_by)
        return participant
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/participants/{participant_id}/edit-history", response_model=List[ParticipantEditHistoryResponse])
def get_participant_edit_history(participant_id: int, db: Session = Depends(get_db)):
    history = db.query(ParticipantEditHistory).filter(
        ParticipantEditHistory.participant_id == participant_id
    ).order_by(ParticipantEditHistory.created_at.desc()).all()
    return history


@router.get("/{application_id}/duplicate-check", response_model=List[DuplicateParticipantWarning])
def check_duplicate_participants(
    application_id: int,
    phone: Optional[str] = Query(None),
    id_number: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    application = db.query(Application).filter(Application.id == application_id).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")
    warnings = app_service.check_duplicate_participants(db, application.group_id, phone, id_number)
    return warnings


@router.post("/{application_id}/participants/import")
def import_participants_excel(
    application_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    import openpyxl
    from io import BytesIO

    application = db.query(Application).filter(Application.id == application_id).first()
    if not application:
        raise HTTPException(status_code=404, detail="Application not found")

    content = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    participants_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        p = ParticipantCreate(
            name=str(row[0]) if row[0] else "",
            gender=str(row[1]) if row[1] else None,
            birth_date=row[2] if row[2] else None,
            phone=str(row[3]) if row[3] else None,
            email=str(row[4]) if row[4] else None,
            address=str(row[5]) if row[5] else None,
            is_leader=bool(row[6]) if row[6] is not None else False,
        )
        participants_data.append(p)

    if not participants_data:
        raise HTTPException(status_code=400, detail="No participant data found in Excel")

    try:
        app = app_service.add_participants(db, application_id, participants_data)
        return {"imported": len(participants_data), "application_id": application_id}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/participants/template")
def download_participant_template():
    import openpyxl
    from fastapi.responses import StreamingResponse
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "参加者信息模板"
    headers = ["姓名", "性别(M/F/O)", "出生日期", "电话", "Email", "联系地址", "是否责任人(True/False)"]
    ws.append(headers)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=participant_template.xlsx"}
    )


@router.post("/{application_id}/cancel", response_model=ApplicationResponse)
def cancel_application(
    application_id: int,
    reason: Optional[str] = Query(None),
    channel: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    try:
        application, refund = app_service.cancel_application(db, application_id, reason, channel)
        return application
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{application_id}/partial-cancel", response_model=ApplicationResponse)
def partial_cancel_application(
    application_id: int,
    data: PartialCancelRequest,
    db: Session = Depends(get_db)
):
    try:
        application, refund = app_service.partial_cancel(
            db, application_id, data.participant_ids, data.reason, data.channel
        )
        return application
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{application_id}/cancel-preview", response_model=CancelPreviewDetailResponse)
def get_cancel_preview(application_id: int, db: Session = Depends(get_db)):
    try:
        preview = app_service.get_cancel_preview(db, application_id)
        return preview
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{application_id}/refunds", response_model=List[RefundDetailResponse])
def get_refunds(application_id: int, db: Session = Depends(get_db)):
    refunds = db.query(Refund).filter(Refund.application_id == application_id).all()
    return refunds


@router.put("/refunds/{refund_id}/approve", response_model=RefundDetailResponse)
def approve_refund(
    refund_id: int,
    data: RefundApprovalRequest,
    db: Session = Depends(get_db)
):
    try:
        refund = app_service.approve_refund(db, refund_id, data.approved, data.approved_by)
        return refund
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{application_id}/reminder-logs", response_model=List[ReminderLogResponse])
def get_reminder_logs(application_id: int, db: Session = Depends(get_db)):
    from models import ReminderLog
    logs = db.query(ReminderLog).filter(ReminderLog.application_id == application_id).all()
    return logs


@router.post("/{application_id}/generate-order-no", response_model=PaymentOrderResponse)
def generate_order_no(
    application_id: int,
    order_type: str = Query("payment_order"),
    db: Session = Depends(get_db)
):
    try:
        order = app_service.generate_payment_order_no(db, application_id, order_type)
        return order
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
